"""
CSV Import Service - Phase 1.3
Handles CSV parsing, validation, column mapping, and bulk import of agents/teams.
"""

from app.database import SessionLocal
from app.models.import_job import ImportJob
from app.models.team import Team
from app.services.team_service import TeamService
from app.services.agent_service import AgentService
from typing import List, Dict, Optional, Tuple
from datetime import datetime
from dataclasses import dataclass
import csv
import re
import logging
import uuid
import os
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass
class AgentData:
    """Data structure for agent information from CSV."""
    agent_name: str
    email: str
    employee_id: Optional[str] = None
    team_name: str = ""


class CSVImportService:
    """Handles CSV parsing, validation, column mapping, and upsert."""
    
    def __init__(self):
        self.team_service = TeamService()
        self.agent_service = AgentService()
    
    def start_import_job(
        self,
        company_id: str,
        file_path: str,
        file_name: str,
        created_by: str
    ) -> ImportJob:
        """
        Create import job record. Kick off background task.
        
        Args:
            company_id: Company ID for the import
            file_path: Path to the uploaded CSV file
            file_name: Original filename
            created_by: User ID who initiated the import
            
        Returns:
            ImportJob instance with status='pending'
        """
        db = SessionLocal()
        try:
            # Detect file format and count total rows (excluding header)
            file_format = self._detect_file_format(file_name)
            rows_total = self._count_rows(file_path, file_format)
            
            # Create import_jobs record with status='pending'
            import_job = ImportJob(
                id=str(uuid.uuid4()),
                company_id=company_id,
                source_type='excel' if file_format == 'xlsx' else 'csv',
                source_platform='n/a',
                status='pending',
                file_name=file_name,
                rows_total=rows_total,
                rows_processed=0,
                rows_failed=0,
                validation_errors=None,
                created_by=created_by
            )
            db.add(import_job)
            db.commit()
            db.refresh(import_job)
            
            logger.info(f"Created import job {import_job.id} for {rows_total} rows")
            return import_job
        except Exception as e:
            db.rollback()
            logger.error(f"Error creating import job: {e}")
            raise
        finally:
            db.close()
    
    def process_import_job(self, job_id: str) -> None:
        """
        Process CSV import. Updates import_jobs record.
        This is a synchronous method that can be called from a background task.
        
        Args:
            job_id: Import job ID to process
        """
        db = SessionLocal()
        try:
            # Get import job
            import_job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            if not import_job:
                raise ValueError(f"Import job {job_id} not found")
            
            # Set status='processing'
            import_job.status = 'processing'
            db.commit()
            
            # Read uploaded file
            file_path = self._get_file_path(import_job.file_name)
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Import file not found: {file_path}")
            
            file_format = self._detect_file_format(import_job.file_name)
            # Parse & validate rows
            rows = self.parse_tabular_file(file_path, file_format)
            
            # Track results
            rows_processed = 0
            rows_failed = 0
            validation_errors = []
            
            # Process each row
            for row_num, row_dict in enumerate(rows, start=2):  # Start at 2 (row 1 is header)
                try:
                    # Validate and map row
                    is_valid, agent_data, error_msg = self.validate_and_map_row(
                        row_dict,
                        import_job.company_id
                    )
                    
                    if not is_valid:
                        rows_failed += 1
                        validation_errors.append({
                            "row_num": row_num,
                            "field": "validation",
                            "error": error_msg,
                            "row_data": row_dict
                        })
                        continue
                    
                    # Upsert agent
                    success, error_msg = self.upsert_agent_from_csv(
                        agent_data,
                        import_job.created_by,
                        import_job.company_id
                    )
                    
                    if success:
                        rows_processed += 1
                    else:
                        rows_failed += 1
                        validation_errors.append({
                            "row_num": row_num,
                            "field": "upsert",
                            "error": error_msg,
                            "row_data": row_dict
                        })
                        
                except Exception as e:
                    rows_failed += 1
                    validation_errors.append({
                        "row_num": row_num,
                        "field": "exception",
                        "error": str(e),
                        "row_data": row_dict
                    })
                    logger.error(f"Error processing row {row_num}: {e}")
            
            # Update import_jobs: status='completed', rows_processed, rows_failed, errors
            import_job.status = 'completed' if rows_failed == 0 else 'completed'  # Still completed even with errors
            import_job.rows_processed = rows_processed
            import_job.rows_failed = rows_failed
            import_job.validation_errors = validation_errors if validation_errors else None
            import_job.completed_at = datetime.utcnow()
            
            db.commit()
            
            logger.info(
                f"Import job {job_id} completed: {rows_processed} processed, "
                f"{rows_failed} failed"
            )
            
        except Exception as e:
            db.rollback()
            logger.error(f"Error processing import job {job_id}: {e}")
            # Update job status to failed
            try:
                import_job.status = 'failed'
                import_job.validation_errors = [{"error": str(e)}]
                db.commit()
            except:
                pass
            raise
        finally:
            db.close()
    
    def parse_tabular_file(self, file_path: str, file_format: str) -> List[Dict[str, str]]:
        """
        Read CSV/XLSX file and return list of row dicts.
        """
        if file_format == 'csv':
            return self._parse_csv_file(file_path)
        if file_format == 'xlsx':
            return self._parse_xlsx_file(file_path)
        raise ValueError(f"Unsupported file format: {file_format}")

    def _parse_csv_file(self, file_path: str) -> List[Dict[str, str]]:
        rows: List[Dict[str, str]] = []
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                required_columns = ['agent_name', 'email', 'team_name']
                actual_columns = [
                    self._normalize_header(col or "") for col in (reader.fieldnames or [])
                ]
                self._validate_required_columns(required_columns, actual_columns)

                for row in reader:
                    if not any(row.values()):
                        continue
                    normalized_row: Dict[str, str] = {}
                    for key, value in row.items():
                        normalized_key = self._normalize_header(key or "")
                        if normalized_key:
                            normalized_row[normalized_key] = (value or "").strip()
                    rows.append(normalized_row)

            logger.info(f"Parsed {len(rows)} rows from CSV file")
            return rows
        except Exception as e:
            logger.error(f"Error parsing CSV file {file_path}: {e}")
            raise

    def _parse_xlsx_file(self, file_path: str) -> List[Dict[str, str]]:
        rows: List[Dict[str, str]] = []
        workbook = None
        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            sheet = workbook.active
            header_row = next(sheet.iter_rows(values_only=True), None)
            if not header_row:
                raise ValueError("Excel file is empty")
            headers = [self._normalize_header(str(cell)) if cell is not None else "" for cell in header_row]
            required_columns = ['agent_name', 'email', 'team_name']
            self._validate_required_columns(required_columns, headers)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Skip rows with no data
                if not row or not any(row):
                    continue
                row_dict: Dict[str, str] = {}
                for header, value in zip(headers, row):
                    if header:
                        row_dict[header] = str(value).strip() if value is not None else ""
                rows.append(row_dict)
            logger.info(f"Parsed {len(rows)} rows from Excel file")
            return rows
        except Exception as e:
            logger.error(f"Error parsing Excel file {file_path}: {e}")
            raise
        finally:
            if workbook:
                workbook.close()

    def _validate_required_columns(self, required: List[str], actual: List[str]) -> None:
        actual_set = {col.strip().lower() for col in actual if col}
        missing = [col for col in required if col not in actual_set]
        if missing:
            raise ValueError(
                f"Missing required columns: {', '.join(missing)}. "
                f"Found columns: {', '.join(actual)}"
            )
    
    def validate_and_map_row(
        self,
        row: Dict[str, str],
        company_id: str
    ) -> Tuple[bool, Optional[AgentData], str]:
        """
        Validate a single CSV row. Return (is_valid, agent_data, error_msg).
        
        Args:
            row: Dictionary representing a CSV row
            company_id: Company ID for team lookup/creation
            
        Returns:
            Tuple of (is_valid: bool, agent_data: Optional[AgentData], error_msg: str)
        """
        # 1. Check required fields: agent_name, email, team_name
        agent_name = row.get('agent_name', '').strip()
        email = row.get('email', '').strip()
        team_name = row.get('team_name', '').strip()
        employee_id = row.get('employee_id', '').strip() or None
        
        if not agent_name:
            return False, None, "agent_name is required"
        
        if not email:
            return False, None, "email is required"
        
        if not team_name:
            return False, None, "team_name is required"
        
        # 2. Validate email format
        if not self._is_valid_email(email):
            return False, None, f"Invalid email format: {email}"
        
        # 3. Check if team_name exists (or create team if not exists)
        # This will be handled in upsert_agent_from_csv to avoid creating teams
        # during validation phase
        
        # 4. Return (True, AgentData(...)) if valid
        agent_data = AgentData(
            agent_name=agent_name,
            email=email,
            employee_id=employee_id,
            team_name=team_name
        )
        
        return True, agent_data, ""
    
    def upsert_agent_from_csv(
        self,
        agent_data: AgentData,
        created_by: str,
        company_id: str
    ) -> Tuple[bool, str]:
        """
        Upsert agent by email. Log all changes. Return (success, error_msg).
        
        Args:
            agent_data: AgentData instance with agent information
            created_by: User ID who created/updated the agent
            company_id: Company ID
            
        Returns:
            Tuple of (success: bool, error_msg: str)
        """
        db = SessionLocal()
        try:
            from app.models.user import User
            
            # 1. Query users WHERE email=? AND deleted_at IS NULL
            existing_agent = db.query(User).filter(
                User.email == agent_data.email,
                User.deleted_at.is_(None)
            ).first()
            
            # 2. Ensure team exists (create if not exists)
            team = db.query(Team).filter(
                Team.company_id == company_id,
                Team.name == agent_data.team_name,
                Team.deleted_at.is_(None)
            ).first()
            
            if not team:
                # Create team if it doesn't exist
                try:
                    team = self.team_service.create_team(
                        company_id=company_id,
                        name=agent_data.team_name,
                        created_by=created_by
                    )
                    logger.info(f"Created team '{agent_data.team_name}' during CSV import")
                except Exception as e:
                    return False, f"Failed to create team '{agent_data.team_name}': {str(e)}"
            
            # 3. If exists: update name, assign to team if different
            if existing_agent:
                # Update agent details if changed
                updated = False
                if existing_agent.full_name != agent_data.agent_name:
                    old_name = existing_agent.full_name
                    existing_agent.full_name = agent_data.agent_name
                    existing_agent.updated_by = created_by
                    existing_agent.updated_at = datetime.utcnow()
                    updated = True
                
                # Check if agent is already in this team
                from app.models.agent_team import AgentTeamMembership
                existing_membership = db.query(AgentTeamMembership).filter(
                    AgentTeamMembership.agent_id == existing_agent.id,
                    AgentTeamMembership.team_id == team.id,
                    AgentTeamMembership.deleted_at.is_(None)
                ).first()
                
                if not existing_membership:
                    # Assign agent to team
                    try:
                        self.agent_service.assign_agent_to_team(
                            agent_id=existing_agent.id,
                            team_id=team.id,
                            created_by=created_by
                        )
                        updated = True
                    except Exception as e:
                        # If already assigned, that's fine
                        if "already assigned" not in str(e).lower():
                            return False, f"Failed to assign agent to team: {str(e)}"
                
                if updated:
                    db.commit()
                
                return True, ""
            
            # 4. If not exists: create new user
            else:
                try:
                    new_agent = self.agent_service.create_agent(
                        company_id=company_id,
                        email=agent_data.email,
                        full_name=agent_data.agent_name,
                        team_id=team.id,
                        created_by=created_by
                    )
                    logger.info(f"Created agent '{agent_data.email}' during CSV import")
                    return True, ""
                except Exception as e:
                    return False, f"Failed to create agent: {str(e)}"
            
        except Exception as e:
            logger.error(f"Error upserting agent from CSV: {e}")
            return False, str(e)
        finally:
            db.close()
    
    def get_import_job_status(self, job_id: str) -> Optional[ImportJob]:
        """
        Get status of import job (for polling from frontend).
        
        Args:
            job_id: Import job ID
            
        Returns:
            ImportJob instance or None if not found
        """
        db = SessionLocal()
        try:
            import_job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            return import_job
        finally:
            db.close()
    
    def _normalize_header(self, header: str) -> str:
        """Normalize column headers to snake_case for consistent lookups."""
        header = header.strip().lower()
        header = re.sub(r'[^a-z0-9]+', '_', header)
        return header.strip('_')

    def _is_valid_email(self, email: str) -> bool:
        """Validate email format using regex."""
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(pattern, email))
    
    def _count_rows(self, file_path: str, file_format: str) -> int:
        """Count total rows in CSV/XLSX file (excluding header)."""
        try:
            if file_format == 'csv':
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    next(reader, None)
                    return sum(1 for row in reader if any(row))
            if file_format == 'xlsx':
                workbook = None
                try:
                    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                    sheet = workbook.active
                    count = 0
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row and any(row):
                            count += 1
                    return count
                finally:
                    if workbook:
                        workbook.close()
        except Exception as e:
            logger.error(f"Error counting rows in file {file_path}: {e}")
        return 0

    def _detect_file_format(self, file_name: str) -> str:
        """Detect file format from extension."""
        _, ext = os.path.splitext(file_name.lower())
        if ext == '.csv':
            return 'csv'
        if ext in ('.xlsx', '.xlsm'):
            return 'xlsx'
        raise ValueError(f"Unsupported file extension '{ext}'. Only CSV and XLSX are supported.")
    
    def _get_file_path(self, file_name: str) -> str:
        """
        Get full file path for uploaded CSV file.
        For now, assumes files are stored in a temp/uploads directory.
        """
        # TODO: Implement proper file storage path resolution
        # For MVP, assume files are in backend/uploads/ directory
        upload_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'uploads')
        os.makedirs(upload_dir, exist_ok=True)
        return os.path.join(upload_dir, file_name)

